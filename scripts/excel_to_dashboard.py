from openpyxl import load_workbook
from datetime import datetime
import os

# 入力ファイルと出力先
excel_path = "dashboard/タスクスケジュールテンプレート.xlsx"
dashboard_md_path = "dashboard/タスクダッシュボード.md"

# Excelファイルを読み込む
wb = load_workbook(excel_path)
ws = wb["個人スケジュール"]

# ヘッダーを取得して列位置を特定
headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

# Markdownテーブルのヘッダー
md_lines = [
    "## 👤 個人スケジュール\n",
    "| No | タスク | 開始日 | 終了日 | タスク属性 | 担当者 | 備考 |",
    "|----|--------|--------|--------|-------------|--------|------|"
]

# データ行をMarkdown形式に変換
for row in ws.iter_rows(min_row=2, values_only=True):
    no = row[headers["No"]]
    task = row[headers["タスク"]]
    start = row[headers["開始日"]]
    end = row[headers["終了日"]]
    kind = row[headers["タスク属性"]]
    owner = row[headers["担当者"]]
    note = row[headers["備考"]]
    md_lines.append(f"| {no} | {task} | {start} | {end} | {kind} | {owner} | {note} |")

# 出力（タスクダッシュボードとして）
with open(dashboard_md_path, "w", encoding="utf-8") as f:
    f.write("# 📋 タスクダッシュボード\n\n")
    f.write("\n".join(md_lines))
    f.write("\n")
